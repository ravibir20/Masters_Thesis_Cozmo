import openpyxl
wb = openpyxl.load_workbook('TStellaULorenzo_alldata_coding.xlsx')

sheet = wb['Sheet1']
# print(sheet['A1'].value)

input = 16
for i in range(4821, 73830, 1):

        sheet.cell(row=i, column=2).value = input

        col1 = sheet.cell(row=i, column=1).value
        if col1 != 0:
            input = col1

wb.save('TStellaULorenzo_alldata_coding_complete.xlsx')
