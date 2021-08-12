import openpyxl
wb = openpyxl.load_workbook('doaction.xlsx')

sheet = wb['Sheet1']
# print(sheet['A1'].value)

for i in range(2, 73777, 1):

        action = sheet.cell(row=i, column=2).value
        if action != 0:
            sheet.cell(row=i, column=1).value = 1
        else:
            sheet.cell(row=i, column=1).value = 0

wb.save('TStellaULorenzo_alldata_final_doaction.xlsx')
