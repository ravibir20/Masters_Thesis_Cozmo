import openpyxl
wb = openpyxl.load_workbook('TBharatUMomo_modelcheckerexcel.xlsx')

sheet = wb['Sheet1']
# print(sheet['A1'].value)

col=7
total = 0
index = 2
totalrepeated = 0
repeated = 0
index_list = []
while index < 17942:
#for i in range(2, 17817, 1):

        input = sheet.cell(row=index, column=col).value

        if input == 1:
            if repeated == 1:
                totalrepeated += 1
            total += 1
            index_list.append(index)
            index += 100
            repeated = 1

        else:
            index += 1
            repeated = 0

print("total: ",total)
print("totalrepeated: ",totalrepeated)

action_match = 0
skip = 0

for ind in index_list:

    for i in range(51):
        checker = sheet.cell(row=(ind+i), column=2).value
        if checker == 1:
            action_match  += 1
            skip = 1
            break

    if skip == 1:
        skip = 0
    else:
        for i in range(1, 51):
            checker = sheet.cell(row=(ind-i), column=2).value
            if checker == 1:
                action_match  += 1
                break



#wb.save('testingcomplete.xlsx')
#action_match  += 1   #since eg one predicts the first one correct
print("actions matched : ",action_match )
